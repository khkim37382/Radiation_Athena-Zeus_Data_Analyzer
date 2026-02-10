from pathlib import Path
import re
import math
import pandas as pd
import datetime as dt

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

try:
    from scipy.stats import chi2  # type: ignore
    _HAVE_SCIPY = True
except Exception:
    chi2 = None
    _HAVE_SCIPY = False

RO_XLSX_PATH = "/Users/jennakronenberg/Desktop/N3_Function_RO.xlsx"

NUM_BLACKBOXES = 8192

FF_PER_SR = [
    4, 4, 1, 4, 1, 1, 2, 2, 2, 2, 1, 2, 2, 2, 1, 1,
    2, 6, 6, 6, 2, 4, 4, 4, 4, 4, 2, 2, 2, 2, 2, 4,
    6, 6, 1, 5, 5, 3, 4, 4, 4, 4, 1, 2, 3, 2, 2, 4,
    4, 1, 4, 4, 1, 2, 3, 3, 1, 1, 1, 2
]
ATHENA_NUM_SRS = len(FF_PER_SR)

FIT_SCALE = 1.0e15
FIT_FACTOR = 0.001


def safe_float(x, default=0.0) -> float:
    try:
        if x is None:
            return default
        if isinstance(x, str) and x.strip() in {"---", ""}:
            return default
        return float(x)
    except Exception:
        return default


def infer_die_category(board: str) -> str:
    b = (board or "").strip().lower()
    if "athena" in b:
        return "Athena"
    if "zeus" in b:
        return "Zeus"
    return "---"


def parse_freq_key(freq_token: str) -> float:
    s = str(freq_token).strip()
    if not s:
        return 0.0
    try:
        if "." in s:
            return float(s)
        n = float(s)
        if n >= 10:
            #return n / 1000.0
            return n
        return n
    except Exception:
        return 0.0


def vdd_to_mV(vdd_token: str) -> int:
    v = safe_float(vdd_token, 0.0)
    if v <= 0:
        return 0
    if v >= 10:
        return int(round(v))
    return int(round(v * 1000.0))


def normalize_board_name(board: str) -> str:
    return (board or "").upper().replace("-", " ").strip()


def poisson_rate_ci_95(k: int, exposure: float) -> tuple[float, float]:
    if exposure <= 0:
        return (0.0, 0.0)

    k = int(k)
    alpha = 0.05

    if _HAVE_SCIPY and chi2 is not None:
        if k == 0:
            lo = 0.0
            hi = 0.5 * chi2.ppf(1 - alpha / 2, 2 * (k + 1)) / exposure
            return (float(lo), float(hi))
        lo = 0.5 * chi2.ppf(alpha / 2, 2 * k) / exposure
        hi = 0.5 * chi2.ppf(1 - alpha / 2, 2 * (k + 1)) / exposure
        return (float(lo), float(hi))

    z = 1.96
    mu = float(k)
    sigma = math.sqrt(mu) if mu > 0 else 1.0
    k_lo = max(0.0, mu - z * sigma)
    k_hi = mu + z * sigma
    return (k_lo / exposure, k_hi / exposure)


class ROData:
    def __init__(self, ro_path: str):
        self.path = Path(ro_path)
        self.wb = None
        if self.path.exists():
            self.wb = load_workbook(self.path, data_only=True)

    def available(self) -> bool:
        return self.wb is not None

    def _find_w8_block(self, board_norm: str):
        ws = self.wb["W8-only"]
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and v.strip().upper() == board_norm.replace(" ", "-"):
                for rr in range(r, min(r + 30, ws.max_row) + 1):
                    row_vals = [ws.cell(rr, c).value for c in range(1, 10)]
                    nums = [x for x in row_vals if isinstance(x, (int, float))]
                    if len(nums) >= 2 and isinstance(ws.cell(rr, 2).value, (int, float)):
                        header_row = rr
                        vdd_start = rr + 1
                        return ws, header_row, vdd_start
        return None

    def _find_non_w8_block(self, sheet_name: str, board_norm: str):
        ws = self.wb[sheet_name]
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 2).value
            if isinstance(v, str) and normalize_board_name(v) == board_norm:
                header_row = r
                vdd_start = r + 1
                return ws, header_row, vdd_start
        return None

    @staticmethod
    def _closest_numeric(target: float, candidates: list[float]) -> float:
        return min(candidates, key=lambda x: abs(x - target)) if candidates else 0.0

    def lookup_actual_freq_mhz(self, die: str, board: str, vdd_mv: int, freq_key: float) -> float:
        if not self.available():
            return 0.0

        board_norm = normalize_board_name(board)
        die_sheet = "Athena" if die == "Athena" else "Zeus" if die == "Zeus" else None
        if die_sheet is None:
            return 0.0

        if "W8" in board_norm:
            blk = self._find_w8_block(board_norm)
            if blk is not None:
                ws, header_row, vdd_start = blk
                headers = {}
                for c in range(2, ws.max_column + 1):
                    hv = ws.cell(header_row, c).value
                    if isinstance(hv, (int, float)):
                        headers[float(hv)] = c
                if not headers:
                    return 0.0

                freq_col = headers.get(freq_key)
                if freq_col is None:
                    nearest = self._closest_numeric(freq_key, list(headers.keys()))
                    freq_col = headers.get(nearest)

                for r in range(vdd_start, min(vdd_start + 10, ws.max_row) + 1):
                    v = ws.cell(r, 1).value
                    if isinstance(v, str) and "MV" in v.upper():
                        mv = int(round(safe_float(v.upper().replace("MV", "").strip(), 0.0)))
                        if mv == vdd_mv:
                            base_mhz = safe_float(ws.cell(r, freq_col).value, 0.0)
                            if(base_mhz < 1): 
                                return base_mhz * 1000
                            return base_mhz * 128.0
                return 0.0

        blk2 = self._find_non_w8_block(die_sheet, board_norm)
        if blk2 is None:
            ws = self.wb[die_sheet]
            for r in range(1, ws.max_row + 1):
                if isinstance(ws.cell(r, 2).value, str) and isinstance(ws.cell(r, 3).value, (int, float)):
                    blk2 = (ws, r, r + 1)
                    break
            if blk2 is None:
                return 0.0

        ws, header_row, vdd_start = blk2
        headers = {}
        for c in range(3, ws.max_column + 1):
            hv = ws.cell(header_row, c).value
            if isinstance(hv, (int, float)):
                headers[float(hv)] = c
        if not headers:
            return 0.0

        freq_col = headers.get(freq_key/1000)
        if freq_col is None:
            nearest = self._closest_numeric(freq_key, list(headers.keys()))
            freq_col = headers.get(nearest)

        for r in range(vdd_start, min(vdd_start + 10, ws.max_row) + 1):
            v = ws.cell(r, 2).value
            if isinstance(v, str) and "MV" in v.upper():
                mv = int(round(safe_float(v.upper().replace("MV", "").strip(), 0.0)))
                if mv == vdd_mv:
                    ghz_val = safe_float(ws.cell(r, freq_col).value, 0.0)
                    if(ghz_val < 1): 
                                return ghz_val * 1000
                    return ghz_val * 128.0

        return 0.0


def get_timestamp(s: str) -> dt.datetime:
    s = str(s).strip()
    for fmt in ("%m/%d/%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%H:%M:%S", "%m/%d/%y %H:%M"):
        try:
            return dt.datetime.strptime(s, fmt)
        except Exception:
            pass
    return dt.datetime.strptime("1:1:1", "%H:%M:%S")



def compute_duration_seconds_from_seu_csv(csv_path: Path) -> float:
    try:
        df = pd.read_csv(csv_path, usecols=[0])
        if df.empty:
            return 0.0
        start = get_timestamp(df.iloc[0, 0])
        end = get_timestamp(df.iloc[-1, 0])
        dur = (end - start).total_seconds()
        return float(dur) if dur > 0 else 0.0
    except Exception:
        try:
            with open(csv_path, "r") as f:
                f.readline()
                first = None
                last = None
                for line in f:
                    parts = line.split(",")
                    if not parts:
                        continue
                    ts = parts[0].strip()
                    if not ts:
                        continue
                    if first is None:
                        first = ts
                    last = ts
            if first is None or last is None:
                return 0.0
            start = get_timestamp(first)
            end = get_timestamp(last)
            dur = (end - start).total_seconds()
            return float(dur) if dur > 0 else 0.0
        except Exception:
            return 0.0


def infer_fluence_from_particle_and_duration(particle: str, duration_s: float) -> float:
    p = (particle or "").upper()
    if "100U" in p:
        return duration_s * 1_000_000.0
    if "10U" in p:
        return duration_s * 100_000.0
    return 0.0


def parse_seu_filename(stem: str) -> dict:
    parts = stem.split("_")
    meta = {
        "angle": "0",
        "temperature": "---",
        "vdd_core": "---",
        "vdd_io": "1.2",
        "die": "---",
        "particle": "---",
        "run": "---",
        "code_used": "LSB",
        "clock_mode": "2.5",
        "frequency": "---",
        "input_bits": "---",
        "fluence": "---",
        "board": "---",
        "actual_freq": 0.0,
    }

    if len(parts) >= 9 and parts[-1].upper() == "SEU":
        meta["run"] = parts[0]
        meta["vdd_core"] = parts[2]
        meta["input_bits"] = parts[3]
        meta["particle"] = parts[4]
        meta["frequency"] = parts[5]
        meta["temperature"] = parts[6]
        meta["board"] = parts[7]
        meta["die"] = infer_die_category(meta["board"])

    return meta


def extract_run_number_from_path(p: Path) -> int:
    m = re.match(r"^(\d+)_", p.stem)
    return int(m.group(1)) if m else 10**9


def build_run_id_from_seu_stem(stem: str) -> str:
    parts = stem.split("_")
    if len(parts) >= 9 and parts[-1].upper() == "SEU":
        return "_".join(parts[0:8])
    return ""


def load_fluence_map_from_run_log(run_log_path: Path) -> dict:
    fluences = {}
    if not run_log_path.exists():
        return fluences

    with open(run_log_path, "r") as infile:
        infile.readline()
        for line in infile:
            temp1 = line.strip().split(",")
            if len(temp1) < 2:
                continue

            temp = temp1[1].split("_") if len(temp1) < 4 else temp1
            if len(temp) < 8:
                continue

            run_id = "_".join(temp[:8])
            try:
                fluences[run_id] = float(temp1[-1])
            except Exception:
                continue

    return fluences


def resolve_fluence(meta: dict, fpath: Path, fluence_map: dict) -> float:
    stem = fpath.stem
    run_id = build_run_id_from_seu_stem(stem)
    if run_id in fluence_map:
        meta["fluence"] = str(fluence_map[run_id])

    fluence = safe_float(meta.get("fluence"), 0.0)
    if fluence <= 0.0:
        duration_s = compute_duration_seconds_from_seu_csv(fpath)
        fluence = infer_fluence_from_particle_and_duration(meta.get("particle", ""), duration_s)
        meta["fluence"] = str(fluence)

    return float(fluence)


def compute_errors_per_sr(csv_path: Path, sr_count: int = ATHENA_NUM_SRS) -> pd.Series:
    df = pd.read_csv(csv_path)

    cols = {c.strip(): c for c in df.columns}
    sr_col = None
    err_col = None

    for candidate in ["SR", "Sr", "Shift Register", "ShiftRegister"]:
        if candidate in cols:
            sr_col = cols[candidate]
            break

    for candidate in ["Error Count", "ErrorCount", "Errors", "Error"]:
        if candidate in cols:
            err_col = cols[candidate]
            break

    if sr_col is None or err_col is None:
        raise ValueError(f"{csv_path.name}: missing SR / Error Count columns")

    df[sr_col] = pd.to_numeric(df[sr_col], errors="coerce")
    df[err_col] = pd.to_numeric(df[err_col], errors="coerce").fillna(0)
    df = df.dropna(subset=[sr_col])

    sr_vals = df[sr_col].astype(int)
    if len(sr_vals) > 0 and sr_vals.min() == 1 and sr_vals.max() == sr_count:
        sr_vals -= 1

    df["_sr0"] = sr_vals
    summed = df.groupby("_sr0")[err_col].sum()

    full = pd.Series(0, index=range(sr_count), dtype="int64")
    for k, v in summed.items():
        if 0 <= int(k) < sr_count:
            full[int(k)] = int(v)

    return full


def compute_actual_frequency_mhz(meta: dict, ro_data: ROData) -> float:
    freq_key = parse_freq_key(meta.get("frequency", "---"))
    if abs(freq_key - 2.5) < 1e-9:
        return 2.5
    if abs(freq_key - 50.0) < 1e-9:
        return 50.0

    if not ro_data.available():
        return 0.0

    die = meta.get("die", "---")
    board = meta.get("board", "---")
    vdd_mv = vdd_to_mV(meta.get("vdd_core", "---"))
    return ro_data.lookup_actual_freq_mhz(die=die, board=board, vdd_mv=vdd_mv, freq_key=freq_key)


def build_summary_excel(folder: str, out_xlsx: str, ro_data: ROData, sr_count: int = ATHENA_NUM_SRS) -> None:
    if sr_count != len(FF_PER_SR):
        raise ValueError("SR count mismatch with FF_PER_SR")

    folder_path = Path(folder)
    fluence_map = load_fluence_map_from_run_log(folder_path / "RUN_LOG.csv")

    seu_files = sorted(
        [p for p in folder_path.iterdir() if p.is_file() and p.name.upper().endswith("_SEU.CSV")],
        key=lambda p: (extract_run_number_from_path(p), p.name)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    pink = PatternFill("solid", fgColor="E7C3C0")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    labels = [
        ("Angle", "Degrees", "angle"),
        ("Temperature", "C", "temperature"),
        ("Vdd_core", "V", "vdd_core"),
        ("Vdd_io", "V", "vdd_io"),
        ("Die #", "", "die"),
        ("Particle", "", "particle"),
        ("Run", "", "run"),
        ("Code Used", "", "code_used"),
        ("Clock Mode", "", "clock_mode"),
        ("Frequency", "", "frequency"),
        ("Input", "", "input_bits"),
        ("Fluence", "", "fluence"),
        ("Actual Freq (MHz)", "", "actual_freq"),
    ]

    for r, (lab, unit, _) in enumerate(labels, start=1):
        ws.cell(row=r, column=1, value=lab).alignment = left
        ws.cell(row=r, column=2, value=unit).alignment = left
        ws.cell(row=r, column=1).font = header_font

    table_start_row = 15

    for i in range(sr_count):
        ws.cell(row=table_start_row + 1 + i, column=1, value=f"SR-{i}")
        ws.cell(row=table_start_row + 1 + i, column=2, value=NUM_BLACKBOXES).alignment = center
        ws.cell(row=table_start_row + 1 + i, column=3, value=FF_PER_SR[i]).alignment = center

    ws.cell(row=table_start_row, column=2, value="# of BlackBoxes").font = header_font
    ws.cell(row=table_start_row, column=3, value="# of FF/BB").font = header_font

    block_start_col = 5
    block_width = 8
    gap = 2

    for idx, fpath in enumerate(seu_files):
        meta = parse_seu_filename(fpath.stem)
        fluence = resolve_fluence(meta, fpath, fluence_map)
        meta["actual_freq"] = compute_actual_frequency_mhz(meta, ro_data)

        errors = compute_errors_per_sr(fpath, sr_count)
        start_col = block_start_col + idx * (block_width + gap)

        ws.column_dimensions[get_column_letter(start_col)].width = 14
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 28
        ws.column_dimensions[get_column_letter(start_col + 2)].width = 14
        ws.column_dimensions[get_column_letter(start_col + 3)].width = 14
        ws.column_dimensions[get_column_letter(start_col + 4)].width = 14

        for r in list(range(1, len(labels) + 1)) + list(range(table_start_row, table_start_row + 1 + sr_count)):
            for c in range(start_col, start_col + block_width):
                ws.cell(row=r, column=c).fill = pink

        for r, (_, _, key) in enumerate(labels, start=1):
            ws.cell(row=r, column=start_col, value=meta.get(key, "---")).alignment = center
            ws.cell(row=r, column=start_col).font = header_font

        ws.cell(row=table_start_row, column=start_col + 0, value="# of Errors").font = header_font
        ws.cell(row=table_start_row, column=start_col + 1, value="CrossSection/FF(cm^2)").font = header_font
        ws.cell(row=table_start_row, column=start_col + 2, value="Lower").font = header_font
        ws.cell(row=table_start_row, column=start_col + 3, value="Upper").font = header_font
        ws.cell(row=table_start_row, column=start_col + 4, value="FIT").font = header_font

        for i in range(sr_count):
            err_i = int(errors.iloc[i])
            ws.cell(row=table_start_row + 1 + i, column=start_col + 0, value=err_i).alignment = center

            total_ffs = NUM_BLACKBOXES * FF_PER_SR[i]
            exposure = fluence * float(total_ffs)

            xs_per_ff = (err_i / exposure) if exposure > 0 else 0.0
            lo, hi = poisson_rate_ci_95(err_i, exposure)

            ws.cell(row=table_start_row + 1 + i, column=start_col + 1, value=xs_per_ff).alignment = center
            ws.cell(row=table_start_row + 1 + i, column=start_col + 2, value=lo).alignment = center
            ws.cell(row=table_start_row + 1 + i, column=start_col + 3, value=hi).alignment = center

            fit = xs_per_ff * FIT_SCALE * FIT_FACTOR
            ws.cell(row=table_start_row + 1 + i, column=start_col + 4, value=fit).alignment = center

    wb.save(out_xlsx)


def build_long_format_excel(folder: str, out_long_xlsx: str, ro_data: ROData, sr_count: int = ATHENA_NUM_SRS) -> None:
    folder_path = Path(folder)
    fluence_map = load_fluence_map_from_run_log(folder_path / "RUN_LOG.csv")

    seu_files = sorted(
        [p for p in folder_path.iterdir() if p.is_file() and p.name.upper().endswith("_SEU.CSV")],
        key=lambda p: (extract_run_number_from_path(p), p.name)
    )

    rows = []
    for fpath in seu_files:
        meta = parse_seu_filename(fpath.stem)
        fluence = resolve_fluence(meta, fpath, fluence_map)
        errors = compute_errors_per_sr(fpath, sr_count)

        parts = fpath.stem.split("_")
        run_id_like = "_".join(parts[0:8]) if len(parts) >= 8 else fpath.stem

        vdd = safe_float(meta.get("vdd_core"), 0.0)
        freq_key = parse_freq_key(meta.get("frequency", 0.0))
        actual_freq_mhz = compute_actual_frequency_mhz(meta, ro_data)

        for sr in range(sr_count):
            err_cnt = int(errors.iloc[sr])
            ff_per_bb = int(FF_PER_SR[sr])
            total_ffs = NUM_BLACKBOXES * ff_per_bb
            exposure = fluence * float(total_ffs)

            cs = (err_cnt / exposure) if exposure > 0 else 0.0
            lo, hi = poisson_rate_ci_95(err_cnt, exposure)
            fit = cs * FIT_SCALE * FIT_FACTOR

            rows.append({
                "id": run_id_like,
                "run": meta.get("run", "---"),
                "vdd": vdd,
                "input": meta.get("input_bits", "---"),
                "ion": meta.get("particle", "---"),
                "temp": meta.get("temperature", "---"),
                "freq": freq_key,
                "brd": meta.get("board", "---"),
                "actual_freq": actual_freq_mhz,
                "fluence": fluence,
                "die": meta.get("die", "---"),
                "# of blackboxes": NUM_BLACKBOXES,
                "# of FF/BB": ff_per_bb,
                "SR_NUM": sr,
                "err_cnt": err_cnt,
                "cs": cs,
                "upper": hi,
                "lower": lo,
                "FIT": fit,
                "source_file": fpath.name,
            })

    df = pd.DataFrame(rows)

    col_order = [
        "id", "run", "vdd", "input", "ion", "temp", "freq", "brd", "actual_freq",
        "fluence", "die", "# of blackboxes", "# of FF/BB", "SR_NUM", "err_cnt",
        "cs", "upper", "lower", "FIT", "source_file"
    ]
    df = df[[c for c in col_order if c in df.columns]]

    with pd.ExcelWriter(out_long_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="LongFormat", index=False)

        ws = writer.book["LongFormat"]
        for j, col in enumerate(df.columns, start=1):
            width = max(10, min(40, len(col) + 2))
            ws.column_dimensions[get_column_letter(j)].width = width

        last_col = get_column_letter(ws.max_column)
        last_row = ws.max_row
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"
        ws.freeze_panes = "A2"


if __name__ == "__main__":
    folder = "/Users/jennakronenberg/Desktop/N3hf"
    out_summary_xlsx = "/Users/jennakronenberg/Desktop/N3hf/NTVsummary.xlsx"
    out_long_xlsx = "/Users/jennakronenberg/Desktop/N3hf/NTVsummary_long.xlsx"

    ro_data = ROData(RO_XLSX_PATH)
    if not ro_data.available():
        print(f"WARNING: RO file not found at {RO_XLSX_PATH}. actual_freq will be 0.0 (except 2.5/50).")

    build_summary_excel(folder, out_summary_xlsx, ro_data)
    build_long_format_excel(folder, out_long_xlsx, ro_data)

    print(f"Saved summary: {out_summary_xlsx}")
    print(f"Saved long format: {out_long_xlsx}")

    if not _HAVE_SCIPY:
        print("Note: SciPy not found -> using approximate CI for upper/lower. Install SciPy for exact bounds: pip install scipy")
